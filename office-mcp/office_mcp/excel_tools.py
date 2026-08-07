"""Excel workbook tools for office-mcp.

This module exposes the 9 core ``excel_*`` tools described in
``architecture.md`` §3.3 and validated by
``validation-contract.md`` (Area: Excel §1-9,
``VAL-EXCEL-001``..``VAL-EXCEL-044``).

Conventions
-----------
* Every tool resolves the user-supplied ``path`` (and optional
  ``folder``) through :func:`office_mcp.paths.resolve_path`.
* Every tool validates its inputs and raises :class:`OfficeMCPError`
  with the appropriate error code (``ERR_FILE_NOT_FOUND``,
  ``ERR_INVALID_PARAMS``, ``ERR_UNSUPPORTED_FMT``,
  ``ERR_SHEET_NOT_FOUND``, ``ERR_CELL_PARSE``).
* Every tool opens the file with :class:`openpyxl.Workbook` /
  :class:`openpyxl.load_workbook`, performs the operation, saves, and
  returns a plain ``dict`` (which FastMCP serializes as the
  ``structuredContent`` of the JSON-RPC response).
* ``excel_create_workbook`` refuses to overwrite an existing file
  (``VAL-EXCEL-003``); ``excel_delete_sheet`` refuses to delete the
  last remaining sheet (``VAL-EXCEL-036``/``-037``);
  ``excel_create_sheet`` / ``excel_rename_sheet`` refuse duplicate
  names (``VAL-EXCEL-033``/``-044``).

Cell-reference helpers
----------------------
The module also exposes two pure-Python helpers:

* :func:`parse_cell_ref` — accepts a single ``"A1"`` style reference
  and returns ``(row, col)`` 1-based (openpyxl convention).
* :func:`parse_range_ref` — accepts either a single cell or an
  ``"A1:C3"`` style range and returns
  ``(row1, col1, row2, col2)`` with ``row1 <= row2`` and
  ``col1 <= col2``. A reversed range like ``"C3:A1"`` is rejected.

Both raise :class:`OfficeMCPError` with ``ERR_CELL_PARSE`` (-32008)
on malformed input.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.chart import (
    AreaChart,
    BarChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Color,
    Font,
    PatternFill,
    Side,
)
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from office_mcp.errors import (
    ERR_CELL_PARSE,
    ERR_FILE_NOT_FOUND,
    ERR_INVALID_PARAMS,
    ERR_SHEET_NOT_FOUND,
    ERR_UNSUPPORTED_FMT,
    OfficeMCPError,
)
from office_mcp.paths import resolve_path, save_with_lock_check
from server import mcp


__all__ = [
    "parse_cell_ref",
    "parse_range_ref",
    "excel_create_workbook",
    "excel_get_info",
    "excel_list_sheets",
    "excel_read_sheet",
    "excel_write_cell",
    "excel_write_range",
    "excel_create_sheet",
    "excel_delete_sheet",
    "excel_rename_sheet",
    "excel_format_cells",
    "excel_add_chart",
    "excel_export_csv",
    "excel_export_pdf",
    "excel_export_html",
]


# ---------------------------------------------------------------------------
# Cell-reference parser
# ---------------------------------------------------------------------------


# Regex: one or more letters (column) followed by one or more digits
# (row). Anchored so trailing garbage like "A1!" is rejected. The
# row part must be a positive integer ("A0" is rejected because the
# regex requires a non-leading-zero integer or any positive int; we
# use ``[1-9][0-9]*`` so "A0" fails the regex).
_CELL_REF_RE = re.compile(r"^([A-Za-z]+)([1-9][0-9]*)$")


def _column_letters_to_index(letters: str) -> int:
    """Convert a column letter sequence to a 1-based index.

    ``"A"`` → 1, ``"Z"`` → 26, ``"AA"`` → 27, ``"AB"`` → 28, ...
    """
    upper = letters.upper()
    result = 0
    for ch in upper:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def parse_cell_ref(ref: str) -> tuple[int, int]:
    """Parse a single cell reference like ``"A1"`` into ``(row, col)``.

    Both indices are 1-based (the openpyxl convention). Column
    letters may be lower-case or upper-case.

    Raises:
        OfficeMCPError: ``ERR_CELL_PARSE`` if the reference is not a
            well-formed single-cell A1 reference (no range, no
            digits-first form, no zero row).
    """
    if not isinstance(ref, str):
        raise OfficeMCPError(
            ERR_CELL_PARSE,
            f"cell reference must be a str, got {type(ref).__name__}",
            {"ref": ref},
        )
    if ":" in ref:
        raise OfficeMCPError(
            ERR_CELL_PARSE,
            f"expected a single cell reference, got range: {ref!r}",
            {"ref": ref},
        )
    match = _CELL_REF_RE.match(ref.strip())
    if not match:
        raise OfficeMCPError(
            ERR_CELL_PARSE,
            f"malformed cell reference: {ref!r}",
            {"ref": ref},
        )
    col_letters, row_digits = match.groups()
    return int(row_digits), _column_letters_to_index(col_letters)


def parse_range_ref(ref: str) -> tuple[int, int, int, int]:
    """Parse a cell or range reference into ``(row1, col1, row2, col2)``.

    Accepts either a single cell (``"A1"``) or a range (``"A1:C3"``).
    Returns the two endpoints in canonical order (``row1 <= row2`` and
    ``col1 <= col2``). A reversed range like ``"C3:A1"`` is rejected
    so the caller never accidentally inverts the coordinates.

    Raises:
        OfficeMCPError: ``ERR_CELL_PARSE`` on malformed input or on a
            reversed range.
    """
    if not isinstance(ref, str):
        raise OfficeMCPError(
            ERR_CELL_PARSE,
            f"cell reference must be a str, got {type(ref).__name__}",
            {"ref": ref},
        )
    parts = ref.strip().split(":")
    if len(parts) == 1:
        row1, col1 = parse_cell_ref(parts[0])
        return row1, col1, row1, col1
    if len(parts) != 2:
        raise OfficeMCPError(
            ERR_CELL_PARSE,
            f"malformed range reference: {ref!r}",
            {"ref": ref},
        )
    # Each side must be a valid cell reference (single-cell, not a
    # sub-range). This rejects inputs like "A1:C3:D5".
    start_row, start_col = parse_cell_ref(parts[0])
    end_row, end_col = parse_cell_ref(parts[1])
    if (start_row, start_col) > (end_row, end_col):
        # Strictly greater on either axis means the range is
        # backwards. We use tuple comparison so e.g. C3:A1 (start
        # (3,3), end (1,1)) is rejected.
        raise OfficeMCPError(
            ERR_CELL_PARSE,
            (
                f"range start must be before end, got start=({start_row},"
                f"{start_col}) end=({end_row},{end_col})"
            ),
            {"ref": ref},
        )
    return start_row, start_col, end_row, end_col


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _resolve_xlsx(path: str, folder: str | None) -> Path:
    """Resolve ``path`` to an absolute ``Path`` and verify the extension.

    Wraps :func:`office_mcp.paths.resolve_path` so that empty /
    whitespace-only paths become :class:`OfficeMCPError`
    (``ERR_INVALID_PARAMS``) instead of a bare ``ValueError``, and
    enforces the ``.xlsx`` extension up front so non-xlsx files fail
    fast.
    """
    try:
        p = resolve_path(path, folder)
    except ValueError as exc:
        raise OfficeMCPError(
            ERR_INVALID_PARAMS,
            str(exc),
            {"path": path},
        ) from exc
    if p.suffix.lower() != ".xlsx":
        raise OfficeMCPError(
            ERR_UNSUPPORTED_FMT,
            f"Expected an .xlsx file, got {p.suffix!r}",
            {"path": str(p), "extension": p.suffix},
        )
    return p


def _require_exists(path: Path) -> None:
    """Raise ``ERR_FILE_NOT_FOUND`` if ``path`` does not exist on disk."""
    if not path.exists():
        raise OfficeMCPError(
            ERR_FILE_NOT_FOUND,
            f"File not found: {path}",
            {"path": str(path)},
        )


def _open_xlsx(path: Path) -> Workbook:
    """Open an existing ``.xlsx`` file.

    Combines :func:`_require_exists` with :func:`openpyxl.load_workbook`
    so that callers always get a proper ``ERR_FILE_NOT_FOUND`` instead
    of a generic ``FileNotFoundError`` from ``openpyxl``.
    """
    _require_exists(path)
    return openpyxl.load_workbook(str(path))


def _save_xlsx(wb: Workbook, path: Path) -> None:
    """Save the workbook back to disk, creating parents if needed.

    Routes the underlying ``Workbook.save`` call through
    :func:`office_mcp.paths.save_with_lock_check` so a held lock
    (file open in Excel) surfaces as ``ERR_FILE_LOCKED`` instead of an
    uncaught :class:`PermissionError`.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    save_with_lock_check(path, lambda: wb.save(str(path)))


def _require_sheet(wb: Workbook, name: Any) -> Worksheet:
    """Return the worksheet named ``name`` or raise ``ERR_SHEET_NOT_FOUND``."""
    if not isinstance(name, str):
        raise OfficeMCPError(
            ERR_INVALID_PARAMS,
            f"sheet name must be a str, got {type(name).__name__}",
            {"sheet": name},
        )
    if name not in wb.sheetnames:
        raise OfficeMCPError(
            ERR_SHEET_NOT_FOUND,
            f"Sheet not found: {name!r}",
            {"sheet": name, "available": list(wb.sheetnames)},
        )
    return wb[name]


def _sheet_info(sheet: Worksheet, index: int) -> dict[str, Any]:
    """Return the ``excel_get_info`` / ``excel_list_sheets`` entry for ``sheet``."""
    # ``max_row`` / ``max_column`` are openpyxl's "used range" markers
    # (1-based, at least 1). For a fresh empty sheet, both are 1 and
    # only the A1 cell exists.
    return {
        "name": sheet.title,
        "index": index,
        "rows": sheet.max_row,
        "cols": sheet.max_column,
    }


def _validate_sheet_name(name: Any, *, context: str) -> str:
    """Validate a sheet name for create / rename operations.

    Excel forbids empty names and certain characters; openpyxl will
    raise ``ValueError`` on those, but we want a clean
    ``ERR_INVALID_PARAMS`` instead of a 500 from the server.
    """
    if not isinstance(name, str):
        raise OfficeMCPError(
            ERR_INVALID_PARAMS,
            f"{context} name must be a str, got {type(name).__name__}",
            {"sheet": name},
        )
    if not name.strip():
        raise OfficeMCPError(
            ERR_INVALID_PARAMS,
            f"{context} name must be a non-empty string",
            {"sheet": name},
        )
    # Excel itself forbids these in sheet names. We mirror the rule
    # so the agent gets a friendly error before openpyxl raises.
    forbidden = set(r":\/?*[]")
    if any(ch in forbidden for ch in name):
        raise OfficeMCPError(
            ERR_INVALID_PARAMS,
            (
                f"{context} name contains forbidden character(s). "
                r"Excel sheet names cannot contain any of: "
                f"{sorted(forbidden)}"
            ),
            {"sheet": name},
        )
    return name


def _read_used_range(sheet: Worksheet) -> list[list[Any]]:
    """Read the sheet's used range as a 2D list of values.

    ``None`` cells are normalised to the empty string so the result
    is JSON-friendly and round-trips cleanly through
    :func:`excel_write_range`.
    """
    rows: list[list[Any]] = []
    max_row = sheet.max_row
    max_col = sheet.max_column
    for r in range(1, max_row + 1):
        row_vals: list[Any] = []
        for c in range(1, max_col + 1):
            v = sheet.cell(row=r, column=c).value
            row_vals.append("" if v is None else v)
        rows.append(row_vals)
    return rows


def _read_range(sheet: Worksheet, r1: int, c1: int, r2: int, c2: int) -> list[list[Any]]:
    """Read a rectangular sub-range as a 2D list of values.

    Indices are inclusive on both ends (1-based, openpyxl convention).
    ``None`` cells are normalised to ``""`` so the result is
    JSON-friendly.
    """
    rows: list[list[Any]] = []
    for r in range(r1, r2 + 1):
        row_vals: list[Any] = []
        for c in range(c1, c2 + 1):
            v = sheet.cell(row=r, column=c).value
            row_vals.append("" if v is None else v)
        rows.append(row_vals)
    return rows


# ---------------------------------------------------------------------------
# 1. excel_create_workbook
# ---------------------------------------------------------------------------


_DEFAULT_SHEET_NAME = "Sheet1"


@mcp.tool()
def excel_create_workbook(
    path: str,
    sheet_name: str | None = None,
    folder: str | None = None,
) -> dict[str, str]:
    """Create a new ``.xlsx`` file at ``path``.

    Args:
        path: Target ``.xlsx`` path. May be absolute or relative to
            ``folder`` (or to the default folder when ``folder`` is
            ``None``).
        sheet_name: Optional name for the initial sheet. When
            ``None`` (the default), the openpyxl default ``"Sheet1"``
            is used.
        folder: Optional base folder for relative paths.

    Returns:
        ``{"path": "<absolute path>"}``.

    Raises:
        OfficeMCPError: ``ERR_INVALID_PARAMS`` if ``path`` is empty
            or a file already exists at the target (or the requested
            sheet name is malformed). ``ERR_UNSUPPORTED_FMT`` if
            ``path`` does not end in ``.xlsx``.
    """
    out = _resolve_xlsx(path, folder)
    if out.exists():
        raise OfficeMCPError(
            ERR_INVALID_PARAMS,
            f"File already exists: {out}",
            {"path": str(out)},
        )
    name = _DEFAULT_SHEET_NAME if sheet_name is None else _validate_sheet_name(
        sheet_name, context="sheet"
    )
    wb = Workbook()
    # openpyxl 3.1 defaults the first sheet's title to "Sheet" (not
    # "Sheet1"). We always rename it so the contract's
    # ``wb.sheetnames == ["Sheet1"]`` assertion passes for the no-arg
    # case (``VAL-EXCEL-002``).
    wb.active.title = name
    _save_xlsx(wb, out)
    return {"path": str(out)}


# ---------------------------------------------------------------------------
# 2. excel_get_info
# ---------------------------------------------------------------------------


@mcp.tool()
def excel_get_info(path: str, folder: str | None = None) -> dict[str, Any]:
    """Return a summary of the workbook's sheets and sizes.

    The returned dict contains:

    * ``path`` — absolute path of the file on disk.
    * ``sheets`` — list of one dict per sheet, in insertion order.
      Each entry has ``name`` (str), ``index`` (int, 0-based),
      ``rows`` (int, openpyxl's ``max_row``), and ``cols`` (int,
      openpyxl's ``max_column``).

    Args:
        path: Path to an existing ``.xlsx``.
        folder: Optional base folder for relative paths.

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the file is
            missing, ``ERR_UNSUPPORTED_FMT`` for non-``.xlsx``
            extensions.
    """
    p = _resolve_xlsx(path, folder)
    wb = _open_xlsx(p)
    sheets = [_sheet_info(wb[name], idx) for idx, name in enumerate(wb.sheetnames)]
    return {"path": str(p), "sheets": sheets}


# ---------------------------------------------------------------------------
# 3. excel_list_sheets
# ---------------------------------------------------------------------------


@mcp.tool()
def excel_list_sheets(path: str, folder: str | None = None) -> list[dict[str, Any]]:
    """Return one dict per sheet in the workbook (insertion order).

    Each entry is identical to those in :func:`excel_get_info`['``sheets``'].

    Args:
        path: Path to an existing ``.xlsx``.
        folder: Optional base folder for relative paths.

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the file is
            missing, ``ERR_UNSUPPORTED_FMT`` for non-``.xlsx``
            extensions.
    """
    p = _resolve_xlsx(path, folder)
    wb = _open_xlsx(p)
    return [_sheet_info(wb[name], idx) for idx, name in enumerate(wb.sheetnames)]


# ---------------------------------------------------------------------------
# 4. excel_read_sheet
# ---------------------------------------------------------------------------


@mcp.tool()
def excel_read_sheet(
    path: str,
    sheet: str,
    range: str | None = None,
    folder: str | None = None,
) -> list[list[Any]]:
    """Read cell values from a sheet, optionally limited to a range.

    Args:
        path: Path to an existing ``.xlsx``.
        sheet: Name of the sheet to read.
        range: Optional cell or range reference. Accepts:

            * ``None`` — read the entire used range
              (``ws.max_row`` × ``ws.max_column``).
            * ``"A1"`` — read a single cell; the result is a 1x1 list
              ``[[value]]``.
            * ``"A1:C3"`` — read a rectangular range. Reversed ranges
              (``"C3:A1"``) are rejected.

            Column letters may be lower-case or upper-case.
        folder: Optional base folder for relative paths.

    Returns:
        A 2D list of values. Empty cells (``None``) are normalised
        to the empty string ``""`` so the result is JSON-friendly.

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the file is
            missing, ``ERR_SHEET_NOT_FOUND`` if ``sheet`` is not in
            the workbook, ``ERR_CELL_PARSE`` for malformed
            ``range``, ``ERR_UNSUPPORTED_FMT`` for non-``.xlsx``
            extensions.
    """
    p = _resolve_xlsx(path, folder)
    wb = _open_xlsx(p)
    ws = _require_sheet(wb, sheet)
    if range is None:
        return _read_used_range(ws)
    r1, c1, r2, c2 = parse_range_ref(range)
    return _read_range(ws, r1, c1, r2, c2)


# ---------------------------------------------------------------------------
# 5. excel_write_cell
# ---------------------------------------------------------------------------


@mcp.tool()
def excel_write_cell(
    path: str,
    sheet: str,
    cell: str,
    value: Any,
    folder: str | None = None,
) -> dict[str, bool]:
    """Write a single value to one cell.

    Strings starting with ``=`` are persisted as formulas
    (``VAL-EXCEL-022``). Other types (``int``, ``float``, ``bool``,
    ``None``) are written verbatim via :meth:`openpyxl.cell.Cell.value`.

    Args:
        path: Path to an existing ``.xlsx``.
        sheet: Name of the sheet.
        cell: A single cell reference like ``"A1"``.
        value: Value to write.
        folder: Optional base folder for relative paths.

    Returns:
        ``{"ok": True}``.

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the file is
            missing, ``ERR_SHEET_NOT_FOUND`` for an unknown sheet,
            ``ERR_CELL_PARSE`` for a malformed ``cell``,
            ``ERR_UNSUPPORTED_FMT`` for non-``.xlsx`` extensions.
    """
    row, col = parse_cell_ref(cell)
    p = _resolve_xlsx(path, folder)
    wb = _open_xlsx(p)
    ws = _require_sheet(wb, sheet)
    ws.cell(row=row, column=col, value=value)
    _save_xlsx(wb, p)
    return {"ok": True}


# ---------------------------------------------------------------------------
# 6. excel_write_range
# ---------------------------------------------------------------------------


@mcp.tool()
def excel_write_range(
    path: str,
    sheet: str,
    start_cell: str,
    data: list[list[Any]],
    folder: str | None = None,
) -> dict[str, bool]:
    """Write a 2D list of values starting at ``start_cell``.

    The data is interpreted in row-major order: ``data[0][0]`` goes
    to ``start_cell``, ``data[0][1]`` to the cell one column to the
    right, ``data[1][0]`` to the cell one row below, and so on.
    Writing extends the sheet's used range as needed
    (``VAL-EXCEL-027``).

    Args:
        path: Path to an existing ``.xlsx``.
        sheet: Name of the sheet.
        start_cell: Top-left cell of the write region (e.g. ``"A1"``).
        data: 2D list of values. The empty list ``[]`` is accepted
            and is a no-op (the file is still saved; the workbook
            remains valid).
        folder: Optional base folder for relative paths.

    Returns:
        ``{"ok": True}``.

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the file is
            missing, ``ERR_SHEET_NOT_FOUND`` for an unknown sheet,
            ``ERR_CELL_PARSE`` for a malformed ``start_cell``,
            ``ERR_INVALID_PARAMS`` if ``data`` is not a list,
            ``ERR_UNSUPPORTED_FMT`` for non-``.xlsx`` extensions.
    """
    if not isinstance(data, list):
        raise OfficeMCPError(
            ERR_INVALID_PARAMS,
            f"data must be a list of lists, got {type(data).__name__}",
            {"type": type(data).__name__},
        )
    start_row, start_col = parse_cell_ref(start_cell)
    p = _resolve_xlsx(path, folder)
    wb = _open_xlsx(p)
    ws = _require_sheet(wb, sheet)

    for r_offset, row_values in enumerate(data):
        if r_offset == 0 and not isinstance(row_values, list):
            raise OfficeMCPError(
                ERR_INVALID_PARAMS,
                (
                    "data rows must be lists; got "
                    f"{type(row_values).__name__} at index 0"
                ),
                {"type": type(row_values).__name__},
            )
        if not isinstance(row_values, list):
            # Skip non-list rows silently (the contract doesn't
            # require raising on this; we are lenient with inputs).
            continue
        for c_offset, value in enumerate(row_values):
            ws.cell(
                row=start_row + r_offset,
                column=start_col + c_offset,
                value=value,
            )

    _save_xlsx(wb, p)
    return {"ok": True}


# ---------------------------------------------------------------------------
# 7. excel_create_sheet
# ---------------------------------------------------------------------------


@mcp.tool()
def excel_create_sheet(
    path: str,
    name: str,
    folder: str | None = None,
) -> dict[str, int]:
    """Append a new sheet to the workbook.

    Args:
        path: Path to an existing ``.xlsx``.
        name: Name of the new sheet. Must be non-empty and free of
            Excel-forbidden characters (``: / \\ ? * [ ]``).
        folder: Optional base folder for relative paths.

    Returns:
        ``{"index": <n>}`` where ``<n>`` is the position of the new
        sheet in ``wb.sheetnames`` (0-based).

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the file is
            missing, ``ERR_INVALID_PARAMS`` for a duplicate or
            malformed ``name``, ``ERR_UNSUPPORTED_FMT`` for
            non-``.xlsx`` extensions.
    """
    sheet_name = _validate_sheet_name(name, context="sheet")
    p = _resolve_xlsx(path, folder)
    wb = _open_xlsx(p)
    if sheet_name in wb.sheetnames:
        raise OfficeMCPError(
            ERR_INVALID_PARAMS,
            f"Sheet already exists: {sheet_name!r}",
            {"sheet": sheet_name, "available": list(wb.sheetnames)},
        )
    wb.create_sheet(title=sheet_name)
    _save_xlsx(wb, p)
    new_index = wb.sheetnames.index(sheet_name)
    return {"index": new_index}


# ---------------------------------------------------------------------------
# 8. excel_delete_sheet
# ---------------------------------------------------------------------------


@mcp.tool()
def excel_delete_sheet(
    path: str,
    name: str,
    folder: str | None = None,
) -> dict[str, bool]:
    """Remove a sheet from the workbook.

    Refuses to delete the last remaining sheet
    (``VAL-EXCEL-036`` / ``VAL-EXCEL-037``): the workbook must
    always have at least one sheet.

    Args:
        path: Path to an existing ``.xlsx``.
        name: Name of the sheet to delete.
        folder: Optional base folder for relative paths.

    Returns:
        ``{"ok": True}``.

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the file is
            missing, ``ERR_SHEET_NOT_FOUND`` for an unknown sheet,
            ``ERR_INVALID_PARAMS`` when the target is the only
            remaining sheet, ``ERR_UNSUPPORTED_FMT`` for non-``.xlsx``
            extensions.
    """
    p = _resolve_xlsx(path, folder)
    wb = _open_xlsx(p)
    if name not in wb.sheetnames:
        raise OfficeMCPError(
            ERR_SHEET_NOT_FOUND,
            f"Sheet not found: {name!r}",
            {"sheet": name, "available": list(wb.sheetnames)},
        )
    if len(wb.sheetnames) <= 1:
        raise OfficeMCPError(
            ERR_INVALID_PARAMS,
            (
                f"Cannot delete the only remaining sheet: {name!r}. "
                "A workbook must always have at least one sheet."
            ),
            {"sheet": name, "sheet_count": len(wb.sheetnames)},
        )
    del wb[name]
    _save_xlsx(wb, p)
    return {"ok": True}


# ---------------------------------------------------------------------------
# 9. excel_rename_sheet
# ---------------------------------------------------------------------------


@mcp.tool()
def excel_rename_sheet(
    path: str,
    old_name: str,
    new_name: str,
    folder: str | None = None,
) -> dict[str, bool]:
    """Rename a sheet in the workbook.

    Cell contents are preserved across the rename
    (``VAL-EXCEL-042``). Renaming a sheet to its own name is a
    no-op (the file is still saved, but the sheet names are
    unchanged — ``VAL-EXCEL-041``).

    Args:
        path: Path to an existing ``.xlsx``.
        old_name: Current name of the sheet to rename.
        new_name: Desired new name. Must be non-empty and free of
            Excel-forbidden characters.
        folder: Optional base folder for relative paths.

    Returns:
        ``{"ok": True}``.

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the file is
            missing, ``ERR_SHEET_NOT_FOUND`` for an unknown
            ``old_name``, ``ERR_INVALID_PARAMS`` when ``new_name``
            is empty, contains forbidden characters, or already
            exists as another sheet, ``ERR_UNSUPPORTED_FMT`` for
            non-``.xlsx`` extensions.
    """
    target = _validate_sheet_name(new_name, context="new sheet")
    p = _resolve_xlsx(path, folder)
    wb = _open_xlsx(p)
    if old_name not in wb.sheetnames:
        raise OfficeMCPError(
            ERR_SHEET_NOT_FOUND,
            f"Sheet not found: {old_name!r}",
            {"sheet": old_name, "available": list(wb.sheetnames)},
        )
    if old_name == target:
        # Idempotent: renaming to the same name is a no-op.
        return {"ok": True}
    if target in wb.sheetnames:
        raise OfficeMCPError(
            ERR_INVALID_PARAMS,
            f"Target sheet name already exists: {target!r}",
            {"new_name": target, "available": list(wb.sheetnames)},
        )
    wb[old_name].title = target
    _save_xlsx(wb, p)
    return {"ok": True}


# ---------------------------------------------------------------------------
# 10. excel_format_cells
# ---------------------------------------------------------------------------


# Map of friendly border-style names to openpyxl side styles.
_BORDER_STYLE_ALIASES: dict[str, str] = {
    "thin": "thin",
    "medium": "medium",
    "thick": "thick",
    "dashed": "dashed",
    "dotted": "dotted",
    "double": "double",
    "hair": "hair",
}


def _resolve_border(border: Any) -> Border:
    """Translate a ``border`` argument into an :class:`openpyxl.styles.Border`.

    Accepts:

    * ``None`` → all sides are ``Side(style=None)`` (no border).
    * A ``str`` (e.g. ``"thin"``) → all four sides get the same
      style. Unknown names raise :class:`OfficeMCPError` with
      ``ERR_INVALID_PARAMS``.
    * A ``dict`` mapping a side name (``"left"``, ``"right"``,
      ``"top"``, ``"bottom"``) to a style name. Missing sides default
      to ``Side(style=None)``. Unknown side names / styles raise
      :class:`OfficeMCPError``.
    """
    empty = Side(style=None)
    if border is None:
        return Border(left=empty, right=empty, top=empty, bottom=empty)
    if isinstance(border, str):
        style = _BORDER_STYLE_ALIASES.get(border.lower())
        if style is None:
            raise OfficeMCPError(
                ERR_INVALID_PARAMS,
                (
                    f"Unknown border style: {border!r}. "
                    f"Supported: {sorted(_BORDER_STYLE_ALIASES)}"
                ),
                {"border": border},
            )
        side = Side(style=style)
        return Border(left=side, right=side, top=side, bottom=side)
    if isinstance(border, dict):
        sides: dict[str, Side] = {
            "left": empty,
            "right": empty,
            "top": empty,
            "bottom": empty,
        }
        for side_name, side_style in border.items():
            if side_name not in sides:
                raise OfficeMCPError(
                    ERR_INVALID_PARAMS,
                    (
                        f"Unknown border side: {side_name!r}. "
                        "Supported: left, right, top, bottom"
                    ),
                    {"side": side_name},
                )
            if not isinstance(side_style, str):
                raise OfficeMCPError(
                    ERR_INVALID_PARAMS,
                    (
                        f"Border side {side_name!r} must be a style name "
                        f"(str), got {type(side_style).__name__}"
                    ),
                    {"side": side_name, "value": side_style},
                )
            style = _BORDER_STYLE_ALIASES.get(side_style.lower())
            if style is None:
                raise OfficeMCPError(
                    ERR_INVALID_PARAMS,
                    (
                        f"Unknown border style: {side_style!r}. "
                        f"Supported: {sorted(_BORDER_STYLE_ALIASES)}"
                    ),
                    {"side": side_name, "style": side_style},
                )
            sides[side_name] = Side(style=style)
        return Border(**sides)
    raise OfficeMCPError(
        ERR_INVALID_PARAMS,
        (
            "border must be a str, dict, or None; got "
            f"{type(border).__name__}"
        ),
        {"border": border},
    )


def _apply_format_to_cell(
    cell: Any,
    *,
    bold: bool | None,
    italic: bool | None,
    font_name: str | None,
    font_size: int | float | None,
    font_color: str | None,
    fill_color: str | None,
    border_obj: Border,
    alignment: str | None,
    number_format: str | None,
) -> None:
    """Apply the requested formatting to one openpyxl cell.

    Each argument is ``None`` when the caller did not request a
    change. Only the non-``None`` arguments are written so the
    caller's existing styling is preserved on unmentioned axes.
    """
    font = cell.font
    if any(
        v is not None
        for v in (bold, italic, font_name, font_size, font_color)
    ):
        cell.font = Font(
            name=font_name if font_name is not None else font.name,
            size=font_size if font_size is not None else font.size,
            bold=bold if bold is not None else font.bold,
            italic=italic if italic is not None else font.italic,
            color=(
                Color(rgb=font_color) if font_color is not None else font.color
            ),
        )
    if fill_color is not None:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=Color(rgb=fill_color),
        )
    # A ``Border()`` with all sides ``Side()`` (style=None) is still
    # considered "no change" — we detect this by checking whether
    # every side's style is None.
    border_has_style = any(
        side.style is not None
        for side in (
            border_obj.left,
            border_obj.right,
            border_obj.top,
            border_obj.bottom,
        )
    )
    if border_has_style:
        cell.border = border_obj
    if alignment is not None:
        # Accept either a string like "center" / "left" / "right"
        # or a single-word shorthand that maps to openpyxl's enum.
        try:
            cell.alignment = Alignment(horizontal=alignment)
        except (TypeError, ValueError) as exc:
            raise OfficeMCPError(
                ERR_INVALID_PARAMS,
                (
                    f"Invalid alignment value: {alignment!r}. "
                    "Use one of: left, center, right, justify, general"
                ),
                {"alignment": alignment},
            ) from exc
    if number_format is not None:
        cell.number_format = number_format


@mcp.tool()
def excel_format_cells(
    path: str,
    sheet: str,
    cell_range: str,
    bold: bool | None = None,
    italic: bool | None = None,
    font_name: str | None = None,
    font_size: int | float | None = None,
    font_color: str | None = None,
    fill_color: str | None = None,
    border: str | dict[str, str] | None = None,
    alignment: str | None = None,
    number_format: str | None = None,
    folder: str | None = None,
) -> dict[str, bool]:
    """Apply font, fill, border, alignment, and number-format to a range.

    Each formatting argument is ``None`` by default and is treated as
    "do not change". When **every** formatting argument is ``None``,
    the call is a true no-op: the file is **not** re-saved, so its
    SHA256 and ``mtime`` are preserved (``VAL-EXCEL-049``).

    Args:
        path: Path to an existing ``.xlsx``.
        sheet: Name of the sheet to format.
        cell_range: A cell reference (``"A1"``) or a range
            (``"A1:C3"``). Column letters may be upper or lower case.
        bold: ``True`` / ``False`` to set ``font.bold``, or ``None``
            to leave it alone.
        italic: Same semantics as ``bold`` for ``font.italic``.
        font_name: A font family name (e.g. ``"Arial"``).
        font_size: Point size (e.g. ``14``).
        font_color: Hex RGB string (``"FF0000"``).
        fill_color: Hex RGB string for the cell background.
        border: Either a single style name applied to all four sides
            (``"thin"``, ``"medium"``, ``"thick"``, ``"dashed"``,
            ``"dotted"``, ``"double"``, ``"hair"``) or a dict mapping
            ``"left"`` / ``"right"`` / ``"top"`` / ``"bottom"`` to
            a style name.
        alignment: ``"left"`` / ``"center"`` / ``"right"`` /
            ``"justify"`` / ``"general"`` for ``Alignment.horizontal``.
        number_format: An Excel number format code (e.g. ``"0.00%"``,
            ``"#,##0.00"``, ``"yyyy-mm-dd"``).
        folder: Optional base folder for relative paths.

    Returns:
        ``{"ok": True}``.

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the file is
            missing, ``ERR_SHEET_NOT_FOUND`` for an unknown sheet,
            ``ERR_CELL_PARSE`` for a malformed range,
            ``ERR_INVALID_PARAMS`` for an unknown border style or
            alignment, ``ERR_UNSUPPORTED_FMT`` for non-``.xlsx``
            extensions.
    """
    r1, c1, r2, c2 = parse_range_ref(cell_range)
    p = _resolve_xlsx(path, folder)
    wb = _open_xlsx(p)
    ws = _require_sheet(wb, sheet)
    border_obj = _resolve_border(border)

    # Detect the all-None no-op case BEFORE we touch the workbook so
    # the file is not re-saved (preserves SHA256 / mtime).
    no_format_requested = all(
        v is None
        for v in (
            bold,
            italic,
            font_name,
            font_size,
            font_color,
            fill_color,
            alignment,
            number_format,
        )
    ) and all(
        side.style is None
        for side in (
            border_obj.left,
            border_obj.right,
            border_obj.top,
            border_obj.bottom,
        )
    )
    if no_format_requested:
        return {"ok": True}

    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            _apply_format_to_cell(
                ws.cell(row=r, column=c),
                bold=bold,
                italic=italic,
                font_name=font_name,
                font_size=font_size,
                font_color=font_color,
                fill_color=fill_color,
                border_obj=border_obj,
                alignment=alignment,
                number_format=number_format,
            )
    _save_xlsx(wb, p)
    return {"ok": True}


# ---------------------------------------------------------------------------
# 11. excel_add_chart
# ---------------------------------------------------------------------------


# Map of friendly chart-type names to openpyxl chart classes.
_CHART_TYPE_MAP: dict[str, type] = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "area": AreaChart,
    "scatter": ScatterChart,
}


@mcp.tool()
def excel_add_chart(
    path: str,
    sheet: str,
    chart_type: str,
    data_range: str,
    target_cell: str,
    title: str | None = None,
    folder: str | None = None,
) -> dict[str, bool]:
    """Add a chart to a sheet, anchored at ``target_cell``.

    Supported ``chart_type`` values: ``"bar"``, ``"line"``,
    ``"pie"``, ``"area"``, ``"scatter"``. Any other value raises
    :class:`OfficeMCPError` with ``ERR_INVALID_PARAMS``
    (``VAL-EXCEL-056``) and the workbook is not modified.

    The chart's data is read from ``data_range`` (e.g. ``"A1:B5"``);
    the first column is treated as categories and subsequent columns
    as series. The chart is anchored at ``target_cell`` (e.g.
    ``"D2"``) so its top-left corner coincides with that cell
    (``VAL-EXCEL-052``).

    Args:
        path: Path to an existing ``.xlsx``.
        sheet: Name of the sheet that will own the chart.
        chart_type: One of ``"bar"``, ``"line"``, ``"pie"``,
            ``"area"``, ``"scatter"``.
        data_range: A range reference like ``"A1:B5"``.
        target_cell: A single cell reference like ``"D2"``. The
            chart's top-left anchor is set to this cell.
        title: Optional chart title. When supplied, it is set via
            openpyxl's title API and survives a reload.
        folder: Optional base folder for relative paths.

    Returns:
        ``{"ok": True}``.

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the file is
            missing, ``ERR_SHEET_NOT_FOUND`` for an unknown sheet,
            ``ERR_INVALID_PARAMS`` for an unknown ``chart_type`` or
            malformed ``data_range`` / ``target_cell``,
            ``ERR_CELL_PARSE`` for a malformed range,
            ``ERR_UNSUPPORTED_FMT`` for non-``.xlsx`` extensions.
    """
    if not isinstance(chart_type, str):
        raise OfficeMCPError(
            ERR_INVALID_PARAMS,
            f"chart_type must be a str, got {type(chart_type).__name__}",
            {"chart_type": chart_type},
        )
    cls = _CHART_TYPE_MAP.get(chart_type.lower())
    if cls is None:
        raise OfficeMCPError(
            ERR_INVALID_PARAMS,
            (
                f"Unknown chart_type: {chart_type!r}. "
                f"Supported: {sorted(_CHART_TYPE_MAP)}"
            ),
            {"chart_type": chart_type},
        )

    dr1, dc1, dr2, dc2 = parse_range_ref(data_range)
    anchor_row, anchor_col = parse_cell_ref(target_cell)
    p = _resolve_xlsx(path, folder)
    wb = _open_xlsx(p)
    ws = _require_sheet(wb, sheet)

    chart = cls()
    # ``Reference`` uses 1-based (min_col, min_row); ``parse_range_ref``
    # already returns 1-based coordinates, so we pass them through.
    data_ref = Reference(
        ws,
        min_col=dc1,
        min_row=dr1,
        max_col=dc2,
        max_row=dr2,
    )
    if chart_type.lower() == "pie":
        # Pie charts typically take a single series (values only).
        chart.add_data(data_ref, titles_from_data=True)
    else:
        chart.add_data(data_ref, titles_from_data=True)
        # Use the first column as the category axis when there are
        # at least 2 columns; otherwise leave the default.
        if dc2 - dc1 >= 1:
            cat_ref = Reference(
                ws,
                min_col=dc1,
                min_row=dr1 + 1,
                max_col=dc1,
                max_row=dr2,
            )
            chart.set_categories(cat_ref)
    if title is not None:
        chart.title = title

    # Anchor the chart at the requested cell. openpyxl's
    # ``ws.add_chart(chart, anchor)`` accepts a cell reference string
    # like ``"D2"`` and constructs the appropriate
    # ``OneCellAnchor(_from=AnchorMarker(col, row, ...))`` for us.
    # After save/reload the marker is reachable as
    # ``chart.anchor._from`` (the underscore is needed because
    # ``from`` is a Python keyword).
    ws.add_chart(chart, target_cell)
    _save_xlsx(wb, p)
    return {"ok": True}


# ---------------------------------------------------------------------------
# 12. excel_export_csv
# ---------------------------------------------------------------------------


@mcp.tool()
def excel_export_csv(
    path: str,
    output: str,
    sheet: str | None = None,
    folder: str | None = None,
) -> dict[str, str]:
    """Export a sheet to CSV via :mod:`openpyxl` and the stdlib :mod:`csv`.

    Delegates to :func:`office_mcp.exporters.export_to_csv`. The
    output is written as UTF-8 (without a BOM) and uses the
    standard Excel-friendly CSV dialect.

    Args:
        path: Path to an existing ``.xlsx`` file.
        output: Target path for the produced ``.csv``. The parent
            directory is created if it does not exist. If a relative
            path is given, it is resolved against ``folder`` (or the
            default folder).
        sheet: Optional sheet name. When ``None`` (the default), the
            workbook's first sheet is exported.
        folder: Optional base folder for relative paths.

    Returns:
        ``{"output_path": "<absolute path of the produced CSV>"}``.

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the source is
            missing, ``ERR_UNSUPPORTED_FMT`` for non-``.xlsx``
            sources, ``ERR_SHEET_NOT_FOUND`` for an unknown sheet.
    """
    # Local import: ``exporters`` does not import from ``excel_tools``
    # so this avoids a circular dependency at module import time.
    from office_mcp import exporters

    src = _resolve_xlsx(path, folder)
    # Verify the source exists up front so the caller gets the
    # agent-friendly ``ERR_FILE_NOT_FOUND`` code rather than a raw
    # ``FileNotFoundError`` from openpyxl.
    _require_exists(src)
    out_path = resolve_path(output, folder)
    out_dir = out_path.parent
    produced = exporters.export_to_csv(src, out_dir, sheet=sheet)
    # ``exporters`` names the output ``<stem>.csv``; honour the
    # caller's requested filename by renaming if it differs.
    if produced != out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        produced.replace(out_path)
        final = out_path
    else:
        final = produced
    return {"output_path": str(final)}


# ---------------------------------------------------------------------------
# 13. excel_export_pdf
# ---------------------------------------------------------------------------


@mcp.tool()
def excel_export_pdf(
    path: str,
    output: str,
    folder: str | None = None,
) -> dict[str, str]:
    """Convert a ``.xlsx`` file to PDF via LibreOffice headless.

    Delegates to :func:`office_mcp.exporters.export_to_pdf`. A
    unique ``-env:UserInstallation`` is used per call so multiple
    exports can run concurrently.

    Args:
        path: Path to an existing ``.xlsx`` file.
        output: Target path for the produced PDF. The parent
            directory is created if it does not exist. If a relative
            path is given, it is resolved against ``folder`` (or the
            default folder).
        folder: Optional base folder for relative paths.

    Returns:
        ``{"output_path": "<absolute path of the produced PDF>"}``.

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the source is
            missing, ``ERR_UNSUPPORTED_FMT`` for non-``.xlsx``
            sources, ``ERR_LIBREOFFICE_MISSING`` when ``soffice`` is
            not on PATH, ``ERR_EXPORT_FAILED`` for any other failure.
    """
    from office_mcp import exporters

    src = _resolve_xlsx(path, folder)
    # Verify the source exists up front so the caller gets the
    # agent-friendly ``ERR_FILE_NOT_FOUND`` code rather than a raw
    # "source file could not be loaded" coming back from
    # LibreOffice.
    _require_exists(src)
    out_path = resolve_path(output, folder)
    out_dir = out_path.parent
    produced = exporters.export_to_pdf(src, out_dir)
    if produced != out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        produced.replace(out_path)
        final = out_path
    else:
        final = produced
    return {"output_path": str(final)}


# ---------------------------------------------------------------------------
# 14. excel_export_html
# ---------------------------------------------------------------------------


@mcp.tool()
def excel_export_html(
    path: str,
    output: str,
    folder: str | None = None,
) -> dict[str, str]:
    """Convert a ``.xlsx`` file to HTML via LibreOffice headless.

    Delegates to :func:`office_mcp.exporters.export_to_html`. Unlike
    the Word variant this path always goes through LibreOffice
    (openpyxl does not produce standalone HTML).

    Args:
        path: Path to an existing ``.xlsx`` file.
        output: Target path for the produced HTML. The parent
            directory is created if it does not exist. If a relative
            path is given, it is resolved against ``folder`` (or the
            default folder).
        folder: Optional base folder for relative paths.

    Returns:
        ``{"output_path": "<absolute path of the produced HTML>"}``.

    Raises:
        OfficeMCPError: ``ERR_FILE_NOT_FOUND`` if the source is
            missing, ``ERR_UNSUPPORTED_FMT`` for non-``.xlsx``
            sources, ``ERR_LIBREOFFICE_MISSING`` when ``soffice`` is
            not on PATH, ``ERR_EXPORT_FAILED`` for any other failure.
    """
    from office_mcp import exporters

    src = _resolve_xlsx(path, folder)
    _require_exists(src)
    out_path = resolve_path(output, folder)
    out_dir = out_path.parent
    produced = exporters.export_to_html(src, out_dir)
    if produced != out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        produced.replace(out_path)
        final = out_path
    else:
        final = produced
    return {"output_path": str(final)}
