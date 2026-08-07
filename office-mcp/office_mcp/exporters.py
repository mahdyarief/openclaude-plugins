"""Export pipeline for office-mcp.

Three exporters, one per output format. All three accept the source
file path and an output directory, and return the absolute ``Path`` of
the produced file.

* :func:`export_to_pdf` — ``.docx`` / ``.xlsx`` / ``.pptx`` → PDF via
  LibreOffice headless.
* :func:`export_to_html` — ``.docx`` → HTML via ``mammoth`` (pure
  Python, no LibreOffice dependency). ``.xlsx`` and ``.pptx`` → HTML
  via LibreOffice headless (not used by Word tools today, but
  provided for completeness / future use).
* :func:`export_to_csv` — ``.xlsx`` → CSV via ``openpyxl`` and the
  stdlib :mod:`csv` module.

Subprocess hygiene
------------------
LibreOffice is invoked via :func:`subprocess.run` with the executable
arguments as a **list** (no shell). Each call uses a unique
``-env:UserInstallation=file:///...`` argument pointing at a fresh
temp directory, which allows concurrent exports to run safely
(``VAL-WORD-079``). The temp directory is **not** removed after the
call — ``soffice`` may still hold file handles to it; the operating
system cleans it up on reboot.

Errors
------
On any failure the exporters raise :class:`OfficeMCPError` with the
appropriate code (``ERR_EXPORT_FAILED`` for runtime issues,
``ERR_LIBREOFFICE_MISSING`` when ``soffice`` is not on PATH, or
``ERR_UNSUPPORTED_FMT`` for unsupported source formats).
"""

from __future__ import annotations

import csv as _csv
import logging
import subprocess
import tempfile
from pathlib import Path

from .config import find_libreoffice
from .errors import (
    ERR_EXPORT_FAILED,
    ERR_LIBREOFFICE_MISSING,
    ERR_UNSUPPORTED_FMT,
    OfficeMCPError,
)

__all__ = [
    "export_to_pdf",
    "export_to_html",
    "export_to_csv",
]

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Subprocess helpers
# ---------------------------------------------------------------------------


def _make_lo_profile_url() -> str:
    """Create a unique LibreOffice ``UserInstallation`` URL.

    Returns a ``file:///`` URL pointing at a fresh temporary directory.
    Passing a different URL per call lets multiple LibreOffice
    invocations run concurrently without contending over the default
    profile lock (``VAL-WORD-079``).
    """
    profile_dir = Path(tempfile.mkdtemp(prefix="lo_"))
    # On Windows, ``os.path.join`` returns a path with backslashes,
    # but the ``file:///`` URL spec requires forward slashes.
    posix_path = profile_dir.as_posix()
    # ``file:///`` URLs need three slashes for an absolute path on
    # POSIX; on Windows the same ``file:///C:/...`` form is accepted
    # by LibreOffice.
    return f"file:///{posix_path.lstrip('/')}"


def _run_libreoffice(
    src: Path,
    out_dir: Path,
    target_format: str,
) -> None:
    """Run ``soffice --headless --convert-to <target_format>``.

    Parameters
    ----------
    src:
        Absolute path to the source file (``docx``/``xlsx``/``pptx``).
    out_dir:
        Directory the produced file is written into. Created if it
        does not exist.
    target_format:
        The convert-to target (``"pdf"`` or ``"html"``).

    Raises
    ------
    OfficeMCPError
        With ``ERR_LIBREOFFICE_MISSING`` if ``soffice`` is not on
        PATH, ``ERR_EXPORT_FAILED`` on a LibreOffice failure or
        timeout, or ``ERR_EXPORT_FAILED`` if the expected output
        file is not produced.
    """
    soffice = find_libreoffice()
    if not soffice:
        raise OfficeMCPError(
            ERR_LIBREOFFICE_MISSING,
            "LibreOffice (soffice) is not installed or not on PATH",
        )
    out_dir.mkdir(parents=True, exist_ok=True)
    profile_url = _make_lo_profile_url()
    cmd = [
        soffice,
        "--headless",
        f"-env:UserInstallation={profile_url}",
        "--convert-to",
        target_format,
        "--outdir",
        str(out_dir),
        str(src),
    ]
    logger.info("libreoffice convert: %s", " ".join(cmd))
    try:
        result = subprocess.run(
            cmd,
            check=True,
            timeout=180,
            capture_output=True,
            text=True,
            # soffice must not inherit our stdin. When the MCP
            # server is launched as a stdio JSON-RPC subprocess its
            # stdin is a pipe; soffice would otherwise try to read
            # from it and hang.
            stdin=subprocess.DEVNULL,
        )
    except subprocess.CalledProcessError as exc:
        # Truncate stderr to avoid sending huge blobs back to the
        # agent.
        stderr = (exc.stderr or "")[:300]
        raise OfficeMCPError(
            ERR_EXPORT_FAILED,
            f"LibreOffice failed: {stderr}",
            {"returncode": exc.returncode, "stderr": stderr},
        ) from exc
    except subprocess.TimeoutExpired as exc:
        raise OfficeMCPError(
            ERR_EXPORT_FAILED,
            "LibreOffice timed out after 180s",
        ) from exc
    # Some LibreOffice builds emit a non-empty stdout on success
    # (e.g. "convert ... using ..."). We do not surface it to the
    # agent, but we log it for diagnostics.
    if result.stdout:
        logger.debug("libreoffice stdout: %s", result.stdout[:300])


# ---------------------------------------------------------------------------
# export_to_pdf
# ---------------------------------------------------------------------------


def export_to_pdf(src: Path, out_dir: Path) -> Path:
    """Convert ``src`` (``.docx`` / ``.xlsx`` / ``.pptx``) to PDF via LibreOffice.

    Parameters
    ----------
    src:
        Source file. Any Office format LibreOffice can read.
    out_dir:
        Directory the produced PDF is written into. Created if it
        does not exist.

    Returns
    -------
    pathlib.Path
        Absolute path of the produced ``.pdf`` file.

    Raises
    ------
    OfficeMCPError
        ``ERR_LIBREOFFICE_MISSING`` if ``soffice`` is not on PATH,
        ``ERR_EXPORT_FAILED`` if LibreOffice fails / times out / does
        not produce a file.
    """
    src = src.resolve()
    out_dir = out_dir.resolve()
    _run_libreoffice(src, out_dir, "pdf")
    pdf = out_dir / (src.stem + ".pdf")
    if not pdf.exists():
        raise OfficeMCPError(
            ERR_EXPORT_FAILED,
            f"PDF not produced: expected {pdf}",
        )
    return pdf


# ---------------------------------------------------------------------------
# export_to_html
# ---------------------------------------------------------------------------


def _docx_to_html_mammoth(src: Path, out_dir: Path) -> Path:
    """Convert a ``.docx`` to HTML using :mod:`mammoth`.

    Mammoth produces an HTML *fragment* (no ``<!doctype>``, no
    ``<html>``/``<head>``/``<body>`` wrappers). We wrap the fragment
    in a minimal HTML5 document so the output is well-formed and the
    validation contract's "<!doctype or <html" check passes.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / (src.stem + ".html")

    with open(src, "rb") as docx_file:
        result = mammoth_convert_to_html(docx_file)
    html_fragment = result.value
    html_doc = (
        "<!doctype html>\n"
        "<html lang=\"en\">\n"
        "<head>\n"
        "<meta charset=\"utf-8\">\n"
        f"<title>{src.stem}</title>\n"
        "</head>\n"
        "<body>\n"
        f"{html_fragment}\n"
        "</body>\n"
        "</html>\n"
    )
    out.write_text(html_doc, encoding="utf-8")
    return out


def export_to_html(src: Path, out_dir: Path) -> Path:
    """Convert ``src`` to HTML.

    * ``.docx`` → :mod:`mammoth` (no LibreOffice dependency).
    * ``.xlsx`` / ``.pptx`` → LibreOffice headless.

    Parameters
    ----------
    src:
        Source file. ``.docx`` is handled by mammoth, other Office
        formats are delegated to LibreOffice.
    out_dir:
        Directory the produced HTML is written into. Created if it
        does not exist.

    Returns
    -------
    pathlib.Path
        Absolute path of the produced ``.html`` file.

    Raises
    ------
    OfficeMCPError
        ``ERR_LIBREOFFICE_MISSING`` if LibreOffice is required and
        not installed, ``ERR_EXPORT_FAILED`` for any other failure.
    """
    src = src.resolve()
    out_dir = out_dir.resolve()
    suffix = src.suffix.lower()
    if suffix == ".docx":
        return _docx_to_html_mammoth(src, out_dir)
    if suffix in (".xlsx", ".pptx"):
        _run_libreoffice(src, out_dir, "html")
        out = out_dir / (src.stem + ".html")
        if not out.exists():
            raise OfficeMCPError(
                ERR_EXPORT_FAILED,
                f"HTML not produced: expected {out}",
            )
        return out
    raise OfficeMCPError(
        ERR_UNSUPPORTED_FMT,
        f"HTML export not supported for {suffix!r}",
        {"extension": suffix},
    )


# ---------------------------------------------------------------------------
# export_to_csv
# ---------------------------------------------------------------------------


def export_to_csv(
    src: Path,
    out_dir: Path,
    sheet: str | None = None,
) -> Path:
    """Convert an ``.xlsx`` workbook to a CSV file.

    Parameters
    ----------
    src:
        Source ``.xlsx`` file.
    out_dir:
        Directory the produced CSV is written into. Created if it
        does not exist.
    sheet:
        Name of the sheet to export. When ``None`` (the default),
        the first worksheet in the workbook is used.

    Returns
    -------
    pathlib.Path
        Absolute path of the produced ``.csv`` file.

    Raises
    ------
    OfficeMCPError
        ``ERR_UNSUPPORTED_FMT`` if ``src`` is not ``.xlsx``,
        ``ERR_SHEET_NOT_FOUND`` (-32007) if the named sheet is not in
        the workbook.
    """
    # Imported lazily so this module has no hard dependency on
    # openpyxl when only the .docx exporters are exercised.
    import openpyxl  # noqa: WPS433 - intentional lazy import

    src = src.resolve()
    out_dir = out_dir.resolve()
    if src.suffix.lower() != ".xlsx":
        raise OfficeMCPError(
            ERR_UNSUPPORTED_FMT,
            f"CSV export is xlsx-only, got {src.suffix!r}",
            {"extension": src.suffix},
        )
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
    if sheet is None:
        target = wb.worksheets[0]
    elif sheet in wb.sheetnames:
        target = wb[sheet]
    else:
        # Local import to avoid a hard dependency in the public
        # module surface (errors.py owns the constant).
        from .errors import ERR_SHEET_NOT_FOUND  # noqa: WPS433

        raise OfficeMCPError(
            ERR_SHEET_NOT_FOUND,
            f"Sheet not found: {sheet!r}",
            {"sheet": sheet, "available": list(wb.sheetnames)},
        )
    out = out_dir / (src.stem + ".csv")
    with open(out, "w", newline="", encoding="utf-8") as fh:
        writer = _csv.writer(fh)
        for row in target.iter_rows(values_only=True):
            writer.writerow("" if v is None else v for v in row)
    return out


# ---------------------------------------------------------------------------
# mammoth helper (lazy import for testability and to surface a clean
# error if mammoth is missing)
# ---------------------------------------------------------------------------


def mammoth_convert_to_html(fileobj):  # type: ignore[no-untyped-def]
    """Thin wrapper around :func:`mammoth.convert_to_html`.

    Defined as a module-level function so tests can patch it via
    ``patch.object(exporters, "mammoth_convert_to_html", ...)``.
    """
    import mammoth  # noqa: WPS433 - intentional lazy import

    return mammoth.convert_to_html(fileobj)
